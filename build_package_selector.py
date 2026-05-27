#!/usr/bin/env python3
"""
Build Package Selector Excel: 2 golf packages at 40% markup
Resonate Club - Trump Premier & Trump Elite
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy
import os

# =============================================================================
# MASTER DATA — Cost → Sell at 40% markup (rounded to whole dollars)
# =============================================================================
MARKUP = 1.4

master = {
    "⛳ Golf Courses": [
        ("Hidden Valley GC",       109, 153),
        ("Sandpiper GC",           130, 182),
        ("Sandpiper GC Replay",      0, 182),  # free 2nd round same day
        ("Trump National LA",      210, 294),
    ],
    "🏨 Hotels (1박/인, Double Occupancy)": [
        ("Beverly Wilshire / Same Tier (5★)", 275, 385),
    ],
    "🚐 Transport": [
        ("Airport Roundtrip",       75, 105),
        ("Driver/Guide + Van (일당)", 125, 175),
    ],
    "🍽️ Meals (1인 1끼)": [
        ("Breakfast (Hotel)",       30,  42),
        ("Lunch",                   40,  56),
        ("Dinner (Regular)",        80, 112),
        ("Exclusive Dinner (Michelin)", 120, 168),
    ],
    "📦 Other": [
        ("Insurance/Tips/Contingency (per day)", 50, 70),
    ],
}

# =============================================================================
# PACKAGE DEFINITIONS
# =============================================================================

class Package:
    def __init__(self, name, code, nights, days):
        self.name = name
        self.code = code
        self.nights = nights
        self.days = days

# --- Package A: Trump Premier (6박7일) ---
pkg_a = Package("트럼프 프리미어", "Trump Premier", 6, 7)

# Golf breakdown for Package A
pkg_a_golf = [
    ("Trump National LA",       1, 210, 294, "1 round"),
    ("Sandpiper GC (Round 1)",  1, 130, 182, "Round 1"),
    ("Sandpiper GC (Replay)",   1,   0, 182, "Free replay — same day"),
    ("Hidden Valley GC",        1, 109, 153, "1 round"),
]

# Non-golf items
pkg_a_hotel     = ("Beverly Wilshire / Same Tier (5★)", 6, 275, 385)
pkg_a_transport = [
    ("Airport Roundtrip",          1,  75, 105),
    ("Driver/Guide + Van (일당)",   7, 125, 175),
]
pkg_a_meals = [
    ("Breakfast (Hotel)",              7, 30,  42),
    ("Lunch",                          6, 40,  56),
    ("Dinner (Regular)",               6, 80, 112),
    ("Exclusive Dinner (Michelin)",    2, 120, 168),
]
pkg_a_insurance = ("Insurance/Tips/Contingency", 7, 50, 70)

# --- Package B: Trump Elite (5박6일) ---
pkg_b = Package("트럼프 엘리트", "Trump Elite", 5, 6)

pkg_b_golf = [
    ("Trump National LA",       2, 210, 294, "2 rounds"),
    ("Hidden Valley GC",        1, 109, 153, "1 round"),
]

pkg_b_hotel     = ("Beverly Wilshire / Same Tier (5★)", 5, 275, 385)
pkg_b_transport = [
    ("Airport Roundtrip",          1,  75, 105),
    ("Driver/Guide + Van (일당)",   6, 125, 175),
]
pkg_b_meals = [
    ("Breakfast (Hotel)",              6, 30,  42),
    ("Lunch",                          5, 40,  56),
    ("Dinner (Regular)",               5, 80, 112),
    ("Exclusive Dinner (Michelin)",    2, 120, 168),
]
pkg_b_insurance = ("Insurance/Tips/Contingency", 6, 50, 70)


# =============================================================================
# STYLING CONSTANTS
# =============================================================================
NAVY      = "1a3450"
GOLD      = "d4af37"
LIGHT_GOLD = "fef9e7"
WHITE     = "FFFFFF"
LIGHT_GRAY = "f5f5f5"
MED_GRAY   = "e8e8e8"
GREEN      = "27ae60"

header_fill   = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_font   = Font(name="Calibri", size=12, bold=True, color=WHITE)
gold_fill     = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
gold_font     = Font(name="Calibri", size=14, bold=True, color=NAVY)
light_gold_fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
cat_fill      = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")
cat_font      = Font(name="Calibri", size=11, bold=True, color=WHITE)
total_fill    = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
total_font    = Font(name="Calibri", size=13, bold=True, color=GOLD)
sub_font      = Font(name="Calibri", size=11, bold=True)
normal_font   = Font(name="Calibri", size=11)
small_font    = Font(name="Calibri", size=10, italic=True)
title_font    = Font(name="Calibri", size=16, bold=True, color=NAVY)
subtitle_font = Font(name="Calibri", size=12, color="555555")
thin_border   = Border(
    left=Side(style="thin", color="d0d0d0"),
    right=Side(style="thin", color="d0d0d0"),
    top=Side(style="thin", color="d0d0d0"),
    bottom=Side(style="thin", color="d0d0d0"),
)
bottom_border = Border(
    bottom=Side(style="medium", color=NAVY),
)
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align    = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_align   = Alignment(horizontal="right", vertical="center")

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, number_format=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:           cell.font = font
    if fill:           cell.fill = fill
    if alignment:      cell.alignment = alignment
    if number_format:  cell.number_format = number_format
    if border:         cell.border = border
    return cell

def write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        set_cell(ws, row, col_start + i, h,
                 font=header_font, fill=header_fill, alignment=center_align)

def write_category_row(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    set_cell(ws, row, col_start, text, font=cat_font, fill=cat_fill, alignment=left_align)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = cat_fill

def write_data_row(ws, row, items, col_start=1, alt=False, num_fmt=None):
    bg = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if alt else None
    for i, val in enumerate(items):
        c = col_start + i
        f = None
        a = left_align if i == 0 else center_align
        nf = None
        if num_fmt and i in num_fmt:
            nf = num_fmt[i]
        set_cell(ws, row, c, val, font=normal_font, fill=bg, alignment=a, number_format=nf, border=thin_border)

def write_subtotal_row(ws, row, col_start, label, value, is_sell=False):
    """Write a subtotal line for a category."""
    set_cell(ws, row, col_start, label, font=sub_font, alignment=left_align)
    if is_sell:
        set_cell(ws, row, col_start + 1, value, font=Font(name="Calibri", size=11, bold=True, color=GOLD),
                 alignment=right_align, number_format='$#,##0')
    else:
        set_cell(ws, row, col_start + 1, value, font=sub_font,
                 alignment=right_align, number_format='$#,##0')

def build_package_line_items(golf_items, hotel, transport_items, meals_items, insurance):
    """
    Returns structured data for a package:
    {
        "golf":           [(desc, cost, sell), ...],
        "golf_cost":      int,
        "golf_sell":      int,
        "hotel":          [(desc, cost, sell), ...],
        "hotel_cost":     int,
        "hotel_sell":     int,
        "transport":      [(desc, cost, sell), ...],
        "transport_cost": int,
        "transport_sell": int,
        "meals":          [(desc, cost, sell), ...],
        "meals_cost":     int,
        "meals_sell":     int,
        "insurance":      [(desc, cost, sell), ...],
        "insurance_cost": int,
        "insurance_sell": int,
        "total_cost":     int,
        "total_sell":     int,
    }
    """
    result = {}

    # Golf
    golf_lines = []
    g_cost = 0
    g_sell = 0
    for name, qty, uc, us, note in golf_items:
        line_cost = uc * qty
        line_sell = us * qty
        g_cost += line_cost
        g_sell += line_sell
        golf_lines.append((f"{name} ×{qty}  ({note})", line_cost, line_sell))
    result["golf"] = golf_lines
    result["golf_cost"] = g_cost
    result["golf_sell"] = g_sell

    # Hotel
    h_name, h_nights, h_uc, h_us = hotel
    h_cost = h_uc * h_nights
    h_sell = h_us * h_nights
    result["hotel"] = [(f"{h_name} ×{h_nights} nights", h_cost, h_sell)]
    result["hotel_cost"] = h_cost
    result["hotel_sell"] = h_sell

    # Transport
    t_lines = []
    t_cost = 0
    t_sell = 0
    for name, qty, uc, us in transport_items:
        line_cost = uc * qty
        line_sell = us * qty
        t_cost += line_cost
        t_sell += line_sell
        t_lines.append((f"{name} ×{qty}", line_cost, line_sell))
    result["transport"] = t_lines
    result["transport_cost"] = t_cost
    result["transport_sell"] = t_sell

    # Meals
    m_lines = []
    m_cost = 0
    m_sell = 0
    for name, qty, uc, us in meals_items:
        line_cost = uc * qty
        line_sell = us * qty
        m_cost += line_cost
        m_sell += line_sell
        m_lines.append((f"{name} ×{qty}", line_cost, line_sell))
    result["meals"] = m_lines
    result["meals_cost"] = m_cost
    result["meals_sell"] = m_sell

    # Insurance
    i_name, i_days, i_uc, i_us = insurance
    i_cost = i_uc * i_days
    i_sell = i_us * i_days
    result["insurance"] = [(f"{i_name} ×{i_days} days", i_cost, i_sell)]
    result["insurance_cost"] = i_cost
    result["insurance_sell"] = i_sell

    result["total_cost"] = g_cost + h_cost + t_cost + m_cost + i_cost
    result["total_sell"] = g_sell + h_sell + t_sell + m_sell + i_sell

    return result


# =============================================================================
# BUILD WORKBOOK
# =============================================================================
wb = openpyxl.Workbook()

# ========================
# SHEET 1: Master Data
# ========================
ws1 = wb.active
ws1.title = "Master Data"

# Title
ws1.merge_cells("A1:E1")
set_cell(ws1, 1, 1, "RESONATE CLUB — Master Data (40% Markup)", font=title_font, alignment=left_align)
set_cell(ws1, 2, 1, "Cost → Sell pricing for all package components", font=subtitle_font, alignment=left_align)

# Headers row 4
headers_md = ["Category", "Item", "Cost (USD)", "Sell (USD)", "Markup %"]
write_header_row(ws1, 4, headers_md)

row = 5
cat_num = 0
for cat_name, items in master.items():
    cat_num += 1
    write_category_row(ws1, row, 1, 5, cat_name)
    row += 1
    alt = False
    for item_name, cost, sell in items:
        markup_pct = f"{(sell/cost - 1)*100:.0f}%" if cost > 0 else "∞ (free replay)"
        write_data_row(ws1, row, [cat_name, item_name, cost, sell, markup_pct],
                       alt=alt, num_fmt={2: '$#,##0', 3: '$#,##0'})
        alt = not alt
        row += 1
    row += 1  # blank row between categories

# Column widths
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 42
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 16

# Freeze panes
ws1.freeze_panes = "A5"

# Auto-filter
ws1.auto_filter.ref = f"A4:E{row - 1}"


# ========================
# SHEET 2: Package Selector
# ========================
ws2 = wb.create_sheet("Package Selector")

# Build computed package data
pkg_a_data = build_package_line_items(pkg_a_golf, pkg_a_hotel, pkg_a_transport, pkg_a_meals, pkg_a_insurance)
pkg_b_data = build_package_line_items(pkg_b_golf, pkg_b_hotel, pkg_b_transport, pkg_b_meals, pkg_b_insurance)

# Title
ws2.merge_cells("A1:H1")
set_cell(ws2, 1, 1, "PACKAGE SELECTOR — Per-Person Pricing Comparison", font=title_font, alignment=left_align)
ws2.merge_cells("A2:H2")
set_cell(ws2, 2, 1, "All prices in USD | Based on double occupancy | 40% markup applied", font=subtitle_font, alignment=left_align)

# Package header rows (row 4)
ws2.merge_cells("A4:D4")
set_cell(ws2, 4, 1, f"PACKAGE A: {pkg_a.name} (Trump Premier)", font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=light_gold_fill, alignment=center_align)
ws2.merge_cells("E4:H4")
set_cell(ws2, 4, 5, f"PACKAGE B: {pkg_b.name} (Trump Elite)", font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=light_gold_fill, alignment=center_align)

# Duration row (row 5)
set_cell(ws2, 5, 1, f"Duration: {pkg_a.nights}박{pkg_a.days}일 ({pkg_a.nights} nights / {pkg_a.days} days)",
         font=small_font, alignment=left_align)
set_cell(ws2, 5, 5, f"Duration: {pkg_b.nights}박{pkg_b.days}일 ({pkg_b.nights} nights / {pkg_b.days} days)",
         font=small_font, alignment=left_align)

# Column headers (row 7)
hdr_a = ["Category / Item", "Qty", "Unit Cost", "Line Cost", "Unit Sell", "Line Sell", "Margin", "Margin %"]
# We'll use 8 columns per package — but only show the important ones
# Simpler layout: Category, Item, Cost, Sell for each package side by side

# Actually, let me do a cleaner layout:
# Col A: Category, Col B: Item, Col C: Cost (A), Col D: Sell (A), Col E: (spacer), Col F: Item, Col G: Cost (B), Col H: Sell (B)

# Let me redesign to be cleaner:
# Cols 1-4: Package A  |  Col 5: spacer  |  Cols 6-9: Package B
# Headers: Category | Item | Cost | Sell

pkg_headers = ["Category", "Item", "Cost", "Sell"]

write_header_row(ws2, 7, pkg_headers, 1)
write_header_row(ws2, 7, pkg_headers, 6)

# ----- Package A rows (cols 1-4) -----
# ----- Package B rows (cols 6-9) -----

def write_package_section(ws, start_row, start_col, pkg_data, pkg_obj):
    """Write a package section starting at start_row, using columns start_col..start_col+3."""
    r = start_row

    categories = [
        ("⛳ Golf",         "golf"),
        ("🏨 Hotel",        "hotel"),
        ("🚐 Transport",    "transport"),
        ("🍽️ Meals",        "meals"),
        ("📦 Insurance",    "insurance"),
    ]

    for cat_label, key in categories:
        write_category_row(ws, r, start_col, start_col + 3, cat_label)
        r += 1
        alt = False
        for desc, cost, sell in pkg_data[key]:
            bg = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if alt else None
            set_cell(ws, r, start_col, cat_label, font=small_font, fill=bg, alignment=left_align, border=thin_border)
            set_cell(ws, r, start_col + 1, desc, font=normal_font, fill=bg, alignment=left_align, border=thin_border)
            set_cell(ws, r, start_col + 2, cost, font=normal_font, fill=bg, alignment=right_align,
                     number_format='$#,##0', border=thin_border)
            set_cell(ws, r, start_col + 3, sell, font=Font(name="Calibri", size=11, color=GOLD), fill=bg,
                     alignment=right_align, number_format='$#,##0', border=thin_border)
            alt = not alt
            r += 1

        # Subtotal
        set_cell(ws, r, start_col + 1, f"Subtotal {cat_label}", font=sub_font, alignment=right_align)
        set_cell(ws, r, start_col + 2, pkg_data[f"{key}_cost"], font=sub_font, alignment=right_align, number_format='$#,##0')
        set_cell(ws, r, start_col + 3, pkg_data[f"{key}_sell"],
                 font=Font(name="Calibri", size=11, bold=True, color=GOLD), alignment=right_align, number_format='$#,##0')
        r += 2

    return r

end_a = write_package_section(ws2, 8, 1, pkg_a_data, pkg_a)
end_b = write_package_section(ws2, 8, 6, pkg_b_data, pkg_b)

# Grand Total row (use max row)
grand_row = max(end_a, end_b)

# Package A Total
set_cell(ws2, grand_row, 1, "🏆 GRAND TOTAL (per person)", font=Font(name="Calibri", size=13, bold=True, color=WHITE),
         fill=total_fill, alignment=left_align)
set_cell(ws2, grand_row, 2, "", fill=total_fill)
set_cell(ws2, grand_row, 3, pkg_a_data["total_cost"], font=total_font, fill=total_fill,
         alignment=right_align, number_format='$#,##0')
set_cell(ws2, grand_row, 4, pkg_a_data["total_sell"], font=total_font, fill=total_fill,
         alignment=right_align, number_format='$#,##0')

# Package B Total
set_cell(ws2, grand_row, 6, "🏆 GRAND TOTAL (per person)", font=Font(name="Calibri", size=13, bold=True, color=WHITE),
         fill=total_fill, alignment=left_align)
set_cell(ws2, grand_row, 7, "", fill=total_fill)
set_cell(ws2, grand_row, 8, pkg_b_data["total_cost"], font=total_font, fill=total_fill,
         alignment=right_align, number_format='$#,##0')
set_cell(ws2, grand_row, 9, pkg_b_data["total_sell"], font=total_font, fill=total_fill,
         alignment=right_align, number_format='$#,##0')

# Column widths
ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 36
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 4   # spacer
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 36
ws2.column_dimensions['H'].width = 15
ws2.column_dimensions['I'].width = 15

ws2.freeze_panes = "A7"


# ========================
# SHEET 3: Package Summary
# ========================
ws3 = wb.create_sheet("Package Summary")

# Title
ws3.merge_cells("A1:F1")
set_cell(ws3, 1, 1, "PACKAGE SUMMARY — Client-Facing Overview", font=title_font, alignment=left_align)
ws3.merge_cells("A2:F2")
set_cell(ws3, 2, 1, "All-inclusive per-person pricing | Resonate Club", font=subtitle_font, alignment=left_align)

# --- Package A Summary ---
r = 4
ws3.merge_cells(f"A{r}:F{r}")
set_cell(ws3, r, 1, f"PACKAGE A: {pkg_a.name}", font=Font(name="Calibri", size=15, bold=True, color=NAVY),
         fill=light_gold_fill, alignment=center_align)
r += 1
set_cell(ws3, r, 1, f"Duration: {pkg_a.nights}박{pkg_a.days}일 ({pkg_a.nights} nights / {pkg_a.days} days)",
         font=sub_font, alignment=left_align)
r += 1
set_cell(ws3, r, 1, "Golf: Trump National LA ×1 | Sandpiper GC ×2 (free replay) | Hidden Valley GC ×1",
         font=normal_font, alignment=left_align)
r += 1
set_cell(ws3, r, 1, f"Hotel: Beverly Wilshire tier, {pkg_a.nights} nights (double occupancy)",
         font=normal_font, alignment=left_align)
r += 1
set_cell(ws3, r, 1, f"Transport: Airport RT + Private Driver/Guide ({pkg_a.days} days)",
         font=normal_font, alignment=left_align)
r += 1
set_cell(ws3, r, 1, f"Meals: Breakfast×{pkg_a.days}, Lunch×{pkg_a.days-1}, Dinner×{pkg_a.days-1}, Exclusive×2",
         font=normal_font, alignment=left_align)
r += 2

# Category summary table for Package A
sum_headers = ["Category", "Cost (USD)", "Sell (USD)", "Margin", "Margin %"]
write_header_row(ws3, r, sum_headers)
r += 1

cats_a = [
    ("⛳ Golf",         pkg_a_data["golf_cost"], pkg_a_data["golf_sell"]),
    ("🏨 Hotel",        pkg_a_data["hotel_cost"], pkg_a_data["hotel_sell"]),
    ("🚐 Transport",    pkg_a_data["transport_cost"], pkg_a_data["transport_sell"]),
    ("🍽️ Meals",        pkg_a_data["meals_cost"], pkg_a_data["meals_sell"]),
    ("📦 Insurance",    pkg_a_data["insurance_cost"], pkg_a_data["insurance_sell"]),
]

alt = False
for label, cost, sell in cats_a:
    margin = sell - cost
    margin_pct = f"{(margin/cost)*100:.1f}%" if cost > 0 else "∞"
    bg = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if alt else None
    set_cell(ws3, r, 1, label, font=sub_font, fill=bg, alignment=left_align, border=thin_border)
    set_cell(ws3, r, 2, cost, font=normal_font, fill=bg, alignment=right_align, number_format='$#,##0', border=thin_border)
    set_cell(ws3, r, 3, sell, font=Font(name="Calibri", size=11, color=GOLD), fill=bg, alignment=right_align,
             number_format='$#,##0', border=thin_border)
    set_cell(ws3, r, 4, margin, font=normal_font, fill=bg, alignment=right_align, number_format='$#,##0', border=thin_border)
    set_cell(ws3, r, 5, margin_pct, font=normal_font, fill=bg, alignment=center_align, border=thin_border)
    alt = not alt
    r += 1

# Grand total A
set_cell(ws3, r, 1, "🏆 TOTAL PER PERSON (Package A)", font=Font(name="Calibri", size=14, bold=True, color=WHITE),
         fill=gold_fill, alignment=left_align)
set_cell(ws3, r, 2, pkg_a_data["total_cost"], font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=gold_fill, alignment=right_align, number_format='$#,##0')
set_cell(ws3, r, 3, pkg_a_data["total_sell"], font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=gold_fill, alignment=right_align, number_format='$#,##0')
total_margin_a = pkg_a_data["total_sell"] - pkg_a_data["total_cost"]
total_margin_pct_a = f"{(total_margin_a/pkg_a_data['total_cost'])*100:.1f}%"
set_cell(ws3, r, 4, total_margin_a, font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=gold_fill, alignment=right_align, number_format='$#,##0')
set_cell(ws3, r, 5, total_margin_pct_a, font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=gold_fill, alignment=center_align)

r += 3

# --- Package B Summary ---
ws3.merge_cells(f"A{r}:F{r}")
set_cell(ws3, r, 1, f"PACKAGE B: {pkg_b.name}", font=Font(name="Calibri", size=15, bold=True, color=NAVY),
         fill=light_gold_fill, alignment=center_align)
r += 1
set_cell(ws3, r, 1, f"Duration: {pkg_b.nights}박{pkg_b.days}일 ({pkg_b.nights} nights / {pkg_b.days} days)",
         font=sub_font, alignment=left_align)
r += 1
set_cell(ws3, r, 1, "Golf: Trump National LA ×2 | Hidden Valley GC ×1",
         font=normal_font, alignment=left_align)
r += 1
set_cell(ws3, r, 1, f"Hotel: Beverly Wilshire tier, {pkg_b.nights} nights (double occupancy)",
         font=normal_font, alignment=left_align)
r += 1
set_cell(ws3, r, 1, f"Transport: Airport RT + Private Driver/Guide ({pkg_b.days} days)",
         font=normal_font, alignment=left_align)
r += 1
set_cell(ws3, r, 1, f"Meals: Breakfast×{pkg_b.days}, Lunch×{pkg_b.days-1}, Dinner×{pkg_b.days-1}, Exclusive×2",
         font=normal_font, alignment=left_align)
r += 2

# Category summary table for Package B
write_header_row(ws3, r, sum_headers)
r += 1

cats_b = [
    ("⛳ Golf",         pkg_b_data["golf_cost"], pkg_b_data["golf_sell"]),
    ("🏨 Hotel",        pkg_b_data["hotel_cost"], pkg_b_data["hotel_sell"]),
    ("🚐 Transport",    pkg_b_data["transport_cost"], pkg_b_data["transport_sell"]),
    ("🍽️ Meals",        pkg_b_data["meals_cost"], pkg_b_data["meals_sell"]),
    ("📦 Insurance",    pkg_b_data["insurance_cost"], pkg_b_data["insurance_sell"]),
]

alt = False
for label, cost, sell in cats_b:
    margin = sell - cost
    margin_pct = f"{(margin/cost)*100:.1f}%" if cost > 0 else "∞"
    bg = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid") if alt else None
    set_cell(ws3, r, 1, label, font=sub_font, fill=bg, alignment=left_align, border=thin_border)
    set_cell(ws3, r, 2, cost, font=normal_font, fill=bg, alignment=right_align, number_format='$#,##0', border=thin_border)
    set_cell(ws3, r, 3, sell, font=Font(name="Calibri", size=11, color=GOLD), fill=bg, alignment=right_align,
             number_format='$#,##0', border=thin_border)
    set_cell(ws3, r, 4, margin, font=normal_font, fill=bg, alignment=right_align, number_format='$#,##0', border=thin_border)
    set_cell(ws3, r, 5, margin_pct, font=normal_font, fill=bg, alignment=center_align, border=thin_border)
    alt = not alt
    r += 1

# Grand total B
set_cell(ws3, r, 1, "🏆 TOTAL PER PERSON (Package B)", font=Font(name="Calibri", size=14, bold=True, color=WHITE),
         fill=gold_fill, alignment=left_align)
set_cell(ws3, r, 2, pkg_b_data["total_cost"], font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=gold_fill, alignment=right_align, number_format='$#,##0')
set_cell(ws3, r, 3, pkg_b_data["total_sell"], font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=gold_fill, alignment=right_align, number_format='$#,##0')
total_margin_b = pkg_b_data["total_sell"] - pkg_b_data["total_cost"]
total_margin_pct_b = f"{(total_margin_b/pkg_b_data['total_cost'])*100:.1f}%"
set_cell(ws3, r, 4, total_margin_b, font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=gold_fill, alignment=right_align, number_format='$#,##0')
set_cell(ws3, r, 5, total_margin_pct_b, font=Font(name="Calibri", size=14, bold=True, color=NAVY),
         fill=gold_fill, alignment=center_align)

# Column widths
ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 4

ws3.freeze_panes = "A4"

# =============================================================================
# PRINT SETTINGS
# =============================================================================
for ws in [ws1, ws2, ws3]:
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

# =============================================================================
# SAVE
# =============================================================================
output_path = os.path.expanduser("~/workspace/resonateclub.github.io/files/Package_Selector_2Packages.xlsx")
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"✅ Saved: {output_path}")

# Copy to secondary location
copy_path = os.path.expanduser("~/workspace/LA_Golf_Packages_2Packages.xlsx")
wb.save(copy_path)
print(f"✅ Copied: {copy_path}")

# =============================================================================
# SUMMARY
# =============================================================================
print()
print("=" * 60)
print("PACKAGE A: 트럼프 프리미어 (Trump Premier) - 6박7일")
print(f"  Per-person Cost:  ${pkg_a_data['total_cost']:>6,}")
print(f"  Per-person Sell:  ${pkg_a_data['total_sell']:>6,}")
print(f"  Margin:           ${total_margin_a:>6,}  ({total_margin_pct_a})")
print()
print("PACKAGE B: 트럼프 엘리트 (Trump Elite) - 5박6일")
print(f"  Per-person Cost:  ${pkg_b_data['total_cost']:>6,}")
print(f"  Per-person Sell:  ${pkg_b_data['total_sell']:>6,}")
print(f"  Margin:           ${total_margin_b:>6,}  ({total_margin_pct_b})")
print("=" * 60)
